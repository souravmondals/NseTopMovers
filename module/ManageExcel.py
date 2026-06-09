import os
import sys
import xlwings as xw
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fyers_apiv3 import fyersModel
from helper import Fyers


def build_fresh_workbook(path, gainers, losers):
    """Create the Excel file from scratch with full formatting (first run only)."""
    wb = Workbook()
 
    def create_table(sheet, title, data, start_row=1):
        headers = ["Symbol", "LTP", "Trade Quantity"]
        keys = ["symbol", "ltp", "trade_quantity"]
        header_fill = PatternFill("solid", start_color="1F4E79")
        alt_fill   = PatternFill("solid", start_color="D6E4F0")
        title_fill = PatternFill("solid", start_color="2E75B6")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
 
        sheet.merge_cells(start_row=start_row, start_column=1,
                          end_row=start_row,   end_column=3)
        tc = sheet.cell(row=start_row, column=1, value=title)
        tc.font = Font(bold=True, color="FFFFFF", size=13, name="Arial")
        tc.fill = title_fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = thin
        sheet.row_dimensions[start_row].height = 22
 
        hr = start_row + 1
        for col, h in enumerate(headers, 1):
            c = sheet.cell(row=hr, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin
        sheet.row_dimensions[hr].height = 18
 
        for i, item in enumerate(data):
            row = start_row + 2 + i
            for col, key in enumerate(keys, 1):
                c = sheet.cell(row=row, column=col, value=item.get(key))
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(
                    horizontal="center" if col > 1 else "left",
                    vertical="center")
                c.border = thin
                if i % 2 == 1:
                    c.fill = alt_fill
                if key == "ltp":
                    c.number_format = "#,##0.00"
                elif key == "trade_quantity":
                    c.number_format = "#,##0"
            sheet.row_dimensions[row].height = 16
 
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 14
        sheet.column_dimensions["C"].width = 18
 
    ws_g = wb.active
    ws_g.title = "Top Gainers"
    create_table(ws_g, "Top Gainers", gainers)
 
    ws_l = wb.create_sheet("Top Losers")
    create_table(ws_l, "Top Losers", losers)
 
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
 
 
def update_live(path, gainers, losers, All_Future_Stocks):
    """
    Write data directly into the already-open Excel workbook via xlwings.
    Excel stays open; no save/close needed.
    """
    # Connect to the open workbook (opens it if not already open)
    try:
        wb = xw.Book(path)
    except Exception:
        xw.App(visible=True).books.open(path)
        wb = xw.Book(path)
 
    def write_dataGainers(sheet_name, data):
        sht = wb.sheets[sheet_name]
        keys = ["symbol", "ltp", "trade_quantity"]
        i=0
        for a, item in enumerate(data):
            #if item.get("symbol") in All_Future_Stocks:
            row = 3 + i  # row 1 = title, row 2 = headers, data starts row 3
            sht.range(f"A{row}").value = item.get("symbol")
            sht.range(f"B{row}").value = item.get("ltp")
            stockDpth = Fyers.get_Stock_Depth(item.get("symbol"))
            sht.range(f"C{row}").value = stockDpth["chp"]
            sht.range(f"D{row}").value = stockDpth["totalbuyqty"]
            sht.range(f"E{row}").value = stockDpth["totalsellqty"]
            i+= 1
    
    def write_dataLosers(sheet_name, data):
        sht = wb.sheets[sheet_name]
        keys = ["symbol", "ltp", "trade_quantity"]
        i=0
        for a, item in enumerate(data):
            #if item.get("symbol") in All_Future_Stocks:
            row = 3 + i  # row 1 = title, row 2 = headers, data starts row 3
            sht.range(f"N{row}").value = item.get("symbol")
            sht.range(f"O{row}").value = item.get("ltp")
            stockDpth = Fyers.get_Stock_Depth(item.get("symbol"))
            sht.range(f"P{row}").value = stockDpth["chp"]
            sht.range(f"Q{row}").value = stockDpth["totalbuyqty"]
            sht.range(f"R{row}").value = stockDpth["totalsellqty"]
            i+= 1
            
 
    write_dataGainers("Data", gainers)
    write_dataLosers("Data",  losers)
 
    # Timestamp in a status cell so you can see the last refresh time   
    sht = wb.sheets["Data"]
    #last_row = sht.range("A3").current_region.last_cell.row
    sht.range(f"K3").value = \
        f"Last updated: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
 
